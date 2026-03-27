import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import math
import re

def compute_gantt():
    # Load data
    df = pd.read_excel('Arena_Costings_TM.xlsx', sheet_name='Revised Budget and Schedule')
    
    # Process rows
    tasks = {}
    
    # Clean Task ID, filter valid
    for index, row in df.iterrows():
        tid = str(row.get('Task ID', '')).strip()
        if tid in ['nan', '.', '', 'None']:
            continue
            
        name = str(row.get('Task Name', ''))
        stream = str(row.get('Workstream', ''))
        deps_raw = str(row.get('Commencement Dependencies', ''))
        if deps_raw in ['nan', 'None']:
            deps_raw = ''
            
        deps = [d.strip() for d in re.split(r'[,;]+', deps_raw) if d.strip()]
        
        start_val = row.get('Start')
        finish_val = row.get('Finish')
        hours = row.get('Project hours / units', 0)
        try:
            hours = float(hours)
            if math.isnan(hours): hours = 0
        except:
            hours = 0
            
        hrs_per_week = row.get('Hours per week avg', 40)
        try:
            hrs_per_week = float(hrs_per_week)
            if math.isnan(hrs_per_week) or hrs_per_week <= 0:
                hrs_per_week = 40
        except:
            hrs_per_week = 40
            
        duration_weeks = max(1, math.ceil(hours / hrs_per_week)) if hours > 0 else 1
        
        # Parse explicit dates
        explicit_start = None
        explicit_finish = None
        if pd.notnull(start_val) and isinstance(start_val, pd.Timestamp):
            explicit_start = start_val
        if pd.notnull(finish_val) and isinstance(finish_val, pd.Timestamp):
            explicit_finish = finish_val
            
        if explicit_start and explicit_finish:
            duration_weeks = max(1, math.ceil((explicit_finish - explicit_start).days / 7.0))
            
        tasks[tid] = {
            'id': tid,
            'name': name,
            'stream': stream,
            'deps': deps,
            'explicit_start': explicit_start,
            'duration_weeks': duration_weeks,
            'start_week': None,
            'end_week': None,
            'resolved': False
        }

    # Find project start date
    valid_starts = [t['explicit_start'] for t in tasks.values() if t['explicit_start']]
    project_start = min(valid_starts) if valid_starts else pd.Timestamp('2026-06-01')

    # Resolve schedule
    max_iters = 100
    for _ in range(max_iters):
        progress = False
        for tid, t in tasks.items():
            if t['resolved']:
                continue
                
            if t['explicit_start']:
                st = max(0, int((t['explicit_start'] - project_start).days / 7))
                t['start_week'] = st
                t['end_week'] = st + t['duration_weeks']
                t['resolved'] = True
                progress = True
                continue
                
            if not t['deps']:
                t['start_week'] = 0
                t['end_week'] = t['duration_weeks']
                t['resolved'] = True
                progress = True
                continue
                
            # Check dependencies
            all_resolved = True
            max_dep_end = 0
            for d in t['deps']:
                if d in tasks:
                    if not tasks[d]['resolved']:
                        all_resolved = False
                        break
                    max_dep_end = max(max_dep_end, tasks[d]['end_week'])
                else:
                    # Dep doesn't exist, ignore
                    pass
                    
            if all_resolved:
                t['start_week'] = max_dep_end
                t['end_week'] = t['start_week'] + t['duration_weeks']
                t['resolved'] = True
                progress = True
                
        if not progress:
            break
            
    # Fallback for unresovled tasks
    for tid, t in tasks.items():
        if not t['resolved']:
            t['start_week'] = 0
            t['end_week'] = t['duration_weeks']
            t['resolved'] = True

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"
    
    headers = ['Task ID', 'Workstream', 'Task Name', 'Dependencies', 'Duration (Wks)']
    max_week = max([t['end_week'] for t in tasks.values()] + [1])
    # Add week columns
    headers.extend([f"W{i+1}" for i in range(max_week)])
    
    ws.append(headers)
    
    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    fill_color = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
    # Helper to sorting IDs correctly, e.g. 1.1, 1.2, ..., 1.10
    def parse_id(tid):
        try:
            parts = [int(p) if p.isdigit() else p for p in tid.split('.')]
            return parts
        except:
            return [tid]

    row_idx = 2
    for tid in sorted(tasks.keys(), key=lambda x: (tasks[x]['start_week'], parse_id(x))):
        t = tasks[tid]
        row_data = [t['id'], t['stream'], t['name'], ', '.join(t['deps']), t['duration_weeks']]
        row_data.extend([''] * max_week)
        ws.append(row_data)
        
        # Color the blocks
        st = t['start_week']
        en = t['end_week']
        for w in range(st, en):
            col_idx = 6 + w # 1-based index, W1 is col 6
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill_color
            if w == st: # First cell in block shows ID
                cell.value = t['id']
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1
        
    # Auto adjust column widths for text
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    # Make week columns smaller
    for w in range(max_week):
        letter = openpyxl.utils.get_column_letter(6 + w)
        ws.column_dimensions[letter].width = 6

    wb.save('Project_Plan_Gantt.xlsx')
    print('Generated Project_Plan_Gantt.xlsx')

if __name__ == '__main__':
    compute_gantt()
